# coding: UTF-8
from schema import Schema
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill, Font

class ExcelBook():
    excelBook = None
    excelSheets = dict()
    resultPath = None

    def __init__(self, resultPath='./result.xlsx'):
        self.excelBook = openpyxl.Workbook(write_only=False)
        self.resultPath = resultPath

    def createExcelSheet(self, index, sheetName):
        self.excelSheets[sheetName] = \
            ExcelSheet(self.excelBook, index, sheetName) 
        return self.excelSheets[sheetName]
    
    def save(self):
        self.excelBook.save(self.resultPath)

class ExcelSheet():
    def __init__(self, excelBook, index, sheetName):
        excelBook.create_sheet(index=index, title=sheetName)
        self.excelSheet = excelBook[sheetName]
        self.excelSchema = {
            "data": dict(),
            "border": dict(),
            "fill": dict(),
            "font": dict(),
            "align": dict(),
            "rowDimension": dict()
        }
 

    def addRowDimension(self, row_num, width):
        self.excelSchema["rowDimension"][row_num] = width

    def addData(self, col, row, text):
        targetCell = col + str(row)
        self.excelSchema['data'][targetCell] = text

    def addBorder(self, cols, rows, style, color, edges):
        side = Side(style=style, color=color)
        for col in cols:
            for row in rows:
                targetCell = col + str(row)
                if targetCell not in self.excelSchema['border']:
                    self.excelSchema['border'][targetCell] = {
                        'left': Side(),
                        'right': Side(),
                        'top': Side(),
                        'bottom': Side()
                    }
                for edge in edges:
                    self.excelSchema['border'][targetCell][edge]=side

    def addBackgroundColor(self, cols, rows, type, color):
        fill = PatternFill(patternType=type, fgColor=color)
        for col in cols:
            for row in rows:
                targetCell = col + str(row)
                self.excelSchema['fill'][targetCell] = fill

    def addAlignment(self, cols, rows, h, v,  
                       wrapText=False):
        align = Alignment(horizontal=h, 
                          vertical=v, 
                          wrapText=wrapText)
        for col in cols:
            for row in rows:
                targetCell = col + str(row)
                self.excelSchema['align'][targetCell] = align
        
    def addFont(self, cols, rows,
                name=None,
                size=None,
                bold=None,
                italic=None,
                color=None):
        font = Font(name=name, 
                    sz=size, 
                    b=bold, 
                    i=italic,
                    color=color)
        for col in cols:
            for row in rows:
                targetCell = col + str(row)
                self.excelSchema['font'][targetCell] = font

    def buildSheet(self):
        for col in self.excelSchema['rowDimension']:
            self.excelSheet.column_dimensions[col].width \
                    = self.excelSchema['rowDimension'][col]
        for cell in self.excelSchema['data'].keys():
            self.excelSheet[cell] = \
                self.excelSchema['data'][cell]
        for cell in self.excelSchema['font'].keys():
            self.excelSheet[cell].font = \
                self.excelSchema['font'][cell]
        for cell in self.excelSchema['fill'].keys():
            self.excelSheet[cell].fill = \
                self.excelSchema['fill'][cell]
        for cell in self.excelSchema['border'].keys():
            side = self.excelSchema['border'][cell]
            border = Border(
                top=side['top'],
                bottom=side['bottom'],
                left=side['left'],
                right=side['right'],
            )
            self.excelSheet[cell].border = border
            


        
if __name__ == "__main__":
    excel = ExcelBook("./result.xlsx")
    sheet = excel.createExcelSheet(0,"testsheet")
    excel.save()